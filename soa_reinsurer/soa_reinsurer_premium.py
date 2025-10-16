import os
import tempfile
import zipfile
import pandas as pd
from datetime import datetime
import re
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# Constants
PREMIUM_COLUMNS = [
    'Reinsurer', 'Address', 'Currency', 'Currency Rate', 'Line', 'Date', 'Our Policy No.', 'Invoice No.', 'Bord Date', 'Inst No.',
    'Due Date', 'Assured', 'Assured Policy No.', 'Binder No.', 'Balance Due', 'Aging', 'REMARKS'
]

# Column width mappings (in pixels converted to Excel width units)
COLUMN_WIDTHS = {
    'Reinsurer': 17.875,
    'Address': 14,
    'Currency': 3.5,
    'Currency Rate': 3.5,
    'Line': 3.5,
    'Date': 9.375,
    'Our Policy No.': 21,
    'Invoice No.': 7.625,
    'Bord Date': 9.375,
    'Inst No.': 2.875,
    'Due Date': 9.375,
    'Assured': 21.5,
    'Assured Policy No.': 19.125,
    'Binder No.': 11.25,
    'Balance Due': 11,
    'Aging': 13.25,
    'REMARKS': 17.125,
    'Updates on': 17.5
}

def make_filename_safe(name: str) -> str:
    """Clean reinsurer name for use in filenames"""
    name = str(name).strip()
    name = re.sub(r'[<>:"/\\|?*]', '', name)
    name = re.sub(r'\s+', ' ', name)
    return name[:100]

def determine_aging_premium(row):
    """Determine aging based on which column contains a numerical value"""
    aging_columns = ['CURRENT', 'OVER 30 DAYS', 'OVER 60 DAYS', 'OVER 90 DAYS', 'OVER 120 DAYS', 'OVER 180 DAYS']
    
    for col in aging_columns:
        if col in row:
            val = row[col]
            if pd.isna(val) or str(val).strip() == '-':
                continue
            
            try:
                num_val = float(str(val).strip().replace(',', ''))
                if num_val != 0:
                    return col
            except (ValueError, TypeError):
                continue
    
    return ''

def apply_premium_formatting_merged(file_path, reinsurer_groups):
    """Apply formatting to merged premium Excel file with separate sections per reinsurer
    
    Args:
        file_path: Path to the Excel file
        reinsurer_groups: List of tuples (reinsurer_name, df_for_reinsurer, address)
    """
    wb = load_workbook(file_path)
    ws = wb.active
    
    # Define styles
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    header_font = Font(bold=True)
    center_align = Alignment(horizontal='center', vertical='center')
    separator_font = Font(bold=True, size=11)
    
    # Clear existing content
    ws.delete_rows(1, ws.max_row)
    
    current_row = 1
    
    # Row 1: Company name
    ws.cell(row=current_row, column=1).value = 'PHILIPPINE FIRST INSURANCE CO. INC'
    ws.cell(row=current_row, column=1).font = Font(bold=True, size=12)
    current_row += 1
    
    # Row 2: Statement of account
    ws.cell(row=current_row, column=1).value = 'STATEMENT OF ACCOUNT'
    ws.cell(row=current_row, column=1).font = Font(bold=True)
    current_row += 1
    
    # Row 3: Date
    today = datetime.now().strftime("AS OF %B %d, %Y").upper()
    ws.cell(row=current_row, column=1).value = today
    ws.cell(row=current_row, column=1).font = Font(bold=True)
    current_row += 1
    
    # Row 4: Empty
    current_row += 1
    
    # Row 5: New Facultative Premium
    ws.cell(row=current_row, column=1).value = 'NEW FACULTATIVE PREMIUM'
    ws.cell(row=current_row, column=1).font = Font(bold=True)
    current_row += 1
    
    # Row 6: Empty
    current_row += 1
    
    # Process each reinsurer section
    for group_idx, (reinsurer_name, reinsurer_df, address) in enumerate(reinsurer_groups):
        if group_idx > 0:
            # Add separator between sections
            current_row += 1
            separator_cell = ws.cell(row=current_row, column=1)
            separator_cell.value = '.' * 100
            separator_cell.font = separator_font
            current_row += 2
        
        # Reinsurer name
        ws.cell(row=current_row, column=1).value = reinsurer_name
        ws.cell(row=current_row, column=1).font = Font(bold=True)
        current_row += 1
        
        # Reinsurer address
        ws.cell(row=current_row, column=1).value = address if address else ''
        ws.cell(row=current_row, column=1).font = Font(size=10)
        current_row += 1
        
        # Empty row
        current_row += 1
        
        # === Create header row for this reinsurer's data ===
        header_row = current_row
        
        # Get columns from first row of reinsurer_df
        for col_idx, col_name in enumerate(reinsurer_df.columns, 1):
            cell = ws.cell(row=header_row, column=col_idx)
            cell.value = col_name
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border
            
            col_letter = chr(64 + col_idx)
            if col_name in COLUMN_WIDTHS:
                ws.column_dimensions[col_letter].width = COLUMN_WIDTHS[col_name]
            else:
                ws.column_dimensions[col_letter].width = 15
        
        current_row += 1
        data_start_row = current_row
        
        # === Write data rows (excluding the total row) ===
        max_col = len(reinsurer_df.columns)
        total_balance = 0
        aging_col = None
        balance_col = None
        
        # Find column indices
        for idx, col_name in enumerate(reinsurer_df.columns, 1):
            if col_name == 'Aging':
                aging_col = idx
            elif col_name == 'Balance Due':
                balance_col = idx
        
        # Separate data rows from total row (detect if last row is all NaN/empty in first column)
        total_row_exists = len(reinsurer_df) > 0 and (pd.isna(reinsurer_df.iloc[-1, 0]) or str(reinsurer_df.iloc[-1, 0]).strip() == '')
        data_rows = reinsurer_df.iloc[:-1] if total_row_exists else reinsurer_df
        
        within_120 = 0
        over_120 = 0
        over_180 = 0
        
        for data_idx, (_, row_data) in enumerate(data_rows.iterrows()):
            for col_idx in range(1, max_col + 1):
                cell = ws.cell(row=current_row, column=col_idx)
                value = row_data.iloc[col_idx - 1]
                col_name = reinsurer_df.columns[col_idx - 1]
                
                # Handle Balance Due specially - ensure it's numeric
                if col_name == 'Balance Due' and pd.notna(value) and value != '':
                    try:
                        num_val = float(str(value).replace(',', ''))
                        cell.value = num_val  # Store as number, not string
                        total_balance += num_val
                        # Don't apply formatting yet - do it after all values are written
                    except (ValueError, TypeError):
                        cell.value = value
                else:
                    if pd.notna(value) and value != '':
                        cell.value = value
                
                cell.border = thin_border
                
                # Track aging for summary
                if col_name == 'Aging' and aging_col and balance_col:
                    aging_val = str(value).strip().upper() if pd.notna(value) else ''
                    balance_val = row_data.iloc[balance_col - 1]
                    
                    try:
                        balance_num = float(str(balance_val).replace(',', ''))
                        if aging_val in ['CURRENT', 'OVER 30 DAYS', 'OVER 60 DAYS', 'OVER 90 DAYS']:
                            within_120 += balance_num
                        elif aging_val == 'OVER 120 DAYS':
                            over_120 += balance_num
                        elif aging_val == 'OVER 180 DAYS':
                            over_180 += balance_num
                    except (ValueError, TypeError):
                        pass
            
            current_row += 1
        
        data_end_row = current_row - 1
        
        # === NOW apply formatting to Balance Due column (after all values written) ===
        if balance_col:
            for row_idx in range(data_start_row, data_end_row + 1):
                cell = ws.cell(row=row_idx, column=balance_col)
                if cell.value is not None and isinstance(cell.value, (int, float)):
                    if cell.value < 0:
                        # Negative: show as (X.XX) in red
                        cell.number_format = '_-* #,##0.00_-;_-* (#,##0.00);_-* "-"??_-;_-@_-'
                        cell.font = Font(color='FF0000')  # Red
                    else:
                        # Positive: standard format
                        cell.number_format = '#,##0.00'
        
        # === Add subtotal row ===
        subtotal_row = current_row
        for col in range(1, max_col + 1):
            cell = ws.cell(row=subtotal_row, column=col)
            cell.border = thin_border
            if col == balance_col:
                cell.value = total_balance  # Store as number
                # Apply format based on sign
                if total_balance < 0:
                    cell.number_format = '_-* #,##0.00_-;_-* (#,##0.00);_-* "-"??_-;_-@_-'
                    cell.font = Font(color='FF0000', bold=True)
                else:
                    cell.number_format = '#,##0.00'
                    cell.font = Font(bold=True)
            else:
                cell.value = ''
        current_row += 2
        
        # === Add aging summary (only non-zero categories) ===
        total_aging = within_120 + over_120 + over_180
        
        ws.cell(row=current_row, column=1).value = 'AGING'
        ws.cell(row=current_row, column=1).font = Font(bold=True)
        current_row += 1
        
        # Only add Within 120 Days if non-zero
        if within_120 != 0:
            ws.cell(row=current_row, column=1).value = 'Within 120 Days - PPW'
            cell_within = ws.cell(row=current_row, column=2)
            cell_within.value = within_120
            cell_within.number_format = '#,##0.00'
            cell_within.font = Font(underline='single')
            current_row += 1
        
        # Only add Over 120 Days if non-zero
        if over_120 != 0:
            ws.cell(row=current_row, column=1).value = 'Over 120 Days'
            cell_over120 = ws.cell(row=current_row, column=2)
            cell_over120.value = over_120
            cell_over120.number_format = '#,##0.00'
            cell_over120.font = Font(underline='single')
            current_row += 1
        
        # Only add Over 180 Days if non-zero
        if over_180 != 0:
            ws.cell(row=current_row, column=1).value = 'Over 180 Days'
            cell_over180 = ws.cell(row=current_row, column=2)
            cell_over180.value = over_180
            cell_over180.number_format = '#,##0.00'
            cell_over180.font = Font(underline='single')
            current_row += 1
        
        total_cell_1 = ws.cell(row=current_row, column=1)
        total_cell_1.value = 'Total'
        total_cell_1.font = Font(bold=True)
        
        total_cell_2 = ws.cell(row=current_row, column=2)
        total_cell_2.value = total_aging
        total_cell_2.font = Font(bold=True, underline='single')
        total_cell_2.number_format = '#,##0.00'
        current_row += 2
    
    # === Add footer at the end (only once) ===
    current_row += 1
    italic_fmt = Font(italic=True)
    bold_fmt = Font(bold=True)
    regular_fmt = Font()
    
    italic_lines = {
        "For your convenience, payments may be made via the BDO Bills Payment facility:"
    }
    bold_lines = {
        "1. BDO Bills Payment", "a. BDO Mobile Application or BDO Web Page", "b. Over the Counter",
        "NOTE: Please make checks payable to PHILIPPINES FIRST INSURANCE CO., INC",
        "      Payments via LBP and BOC Peso are available only through special arrangement via fund transfer, with an advance copy of the remittance schedule."
    }
    footer_lines = [
        "For your convenience, payments may be made via the BDO Bills Payment facility:",
        "", "1. BDO Bills Payment", "a. BDO Mobile Application or BDO Web Page",
        "   i. Biller: Philippines First Insurance Co., Inc.", "   ii. Reference Number: HO-0001",
        "b. Over the Counter", "   i. Company Name: Philippines First Insurance Co., Inc.",
        "   ii. Subscriber Name: Your Company Name", "   iii. Subscriber Account Number: HO-0001",
        "", "NOTE: Please make checks payable to PHILIPPINES FIRST INSURANCE CO., INC",
        "      Payments via LBP and BOC Peso are available only through special arrangement via fund transfer, with an advance copy of the remittance schedule."
    ]
    
    for line in footer_lines:
        cell = ws.cell(row=current_row, column=1)
        cell.value = line
        if line in italic_lines:
            cell.font = italic_fmt
        elif line in bold_lines:
            cell.font = bold_fmt
        else:
            cell.font = regular_fmt
        current_row += 1
    
    wb.save(file_path)

def apply_premium_formatting(file_path, reinsurer_names, is_merged=False):
    """Apply formatting to premium Excel file (for single reinsurer)"""
    wb = load_workbook(file_path)
    ws = wb.active
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    header_font = Font(bold=True)
    center_align = Alignment(horizontal='center', vertical='center')
    
    ws.insert_rows(1, 9)
    header_row = 10

    for col_idx, cell in enumerate(ws[header_row], 1):
        if cell.value:
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border

    ws['A1'] = 'PHILIPPINE FIRST INSURANCE CO. INC'
    ws['A1'].font = Font(bold=False, size=12)
    
    ws['A2'] = 'STATEMENT OF ACCOUNT'
    ws['A2'].font = Font(bold=False)
    
    today = datetime.now().strftime("AS OF %B %d, %Y").upper()
    ws['A3'] = today
    ws['A3'].font = Font(bold=True)
    
    ws['A5'] = 'NEW FACULTATIVE PREMIUM'
    ws['A5'].font = Font(bold=True)
    
    if isinstance(reinsurer_names, list):
        ws['A7'] = reinsurer_names[0]
    else:
        ws['A7'] = reinsurer_names
    ws['A7'].font = Font(bold=True)
    
    address = ''
    address_col = None
    for idx, cell in enumerate(ws[header_row], start=1):
        if cell.value == 'Address':
            address_col = idx
            break
    
    if address_col and ws.max_row > header_row:
        address_value = ws.cell(row=header_row + 1, column=address_col).value
        if address_value:
            address = str(address_value)
    ws['A8'] = address
    ws['A8'].font = Font(size=10)

    data_end_row = header_row + 1
    for row in range(header_row + 1, ws.max_row + 1):
        if ws.cell(row=row, column=1).value:
            data_end_row = row
        else:
            break
    
    subtotal_row = data_end_row + 1
    
    if data_end_row > header_row:
        max_col = ws.max_column
        for row in range(header_row + 1, data_end_row + 1):
            for col in range(1, max_col + 1):
                ws.cell(row=row, column=col).border = thin_border
        
        aging_col = None
        balance_col = None
        
        for idx, cell in enumerate(ws[header_row], start=1):
            if cell.value == 'Aging':
                aging_col = idx
            elif cell.value == 'Balance Due':
                balance_col = idx
        
        within_120 = 0
        over_120 = 0
        over_180 = 0
        total_balance = 0
        
        if aging_col and balance_col:
            for row in range(header_row + 1, data_end_row + 1):
                aging_val = ws.cell(row=row, column=aging_col).value
                balance_cell = ws.cell(row=row, column=balance_col)
                balance_cell_val = balance_cell.value
                
                # Ensure Balance Due is stored as numeric
                if balance_cell_val is not None:
                    try:
                        balance_val = float(str(balance_cell_val).replace(',', ''))
                        balance_cell.value = balance_val  # Ensure it's numeric
                        total_balance += balance_val
                        
                        if aging_val:
                            aging_val = str(aging_val).strip().upper()
                        
                        if aging_val in ['CURRENT', 'OVER 30 DAYS', 'OVER 60 DAYS', 'OVER 90 DAYS']:
                            within_120 += balance_val
                        elif aging_val == 'OVER 120 DAYS':
                            over_120 += balance_val
                        elif aging_val == 'OVER 180 DAYS':
                            over_180 += balance_val
                    except (ValueError, TypeError):
                        continue
        
        # === Apply formatting to Balance Due cells AFTER ensuring they're numeric ===
        if balance_col:
            for row in range(header_row + 1, data_end_row + 1):
                cell = ws.cell(row=row, column=balance_col)
                if cell.value is not None and isinstance(cell.value, (int, float)):
                    if cell.value < 0:
                        # Negative: show as (X.XX) in red
                        cell.number_format = '_-* #,##0.00_-;_-* (#,##0.00);_-* "-"??_-;_-@_-'
                        cell.font = Font(color='FF0000')
                    else:
                        # Positive: standard format
                        cell.number_format = '#,##0.00'
        
        total_aging = within_120 + over_120 + over_180
        
        for col in range(1, max_col + 1):
            cell = ws.cell(row=subtotal_row, column=col)
            cell.border = thin_border
            if col == balance_col:
                cell.value = total_balance  # Numeric value
                if total_balance < 0:
                    cell.number_format = '_-* #,##0.00_-;_-* (#,##0.00);_-* "-"??_-;_-@_-'
                    cell.font = Font(color='FF0000', bold=True)
                else:
                    cell.number_format = '#,##0.00'
                    cell.font = Font(bold=True)
            else:
                cell.value = ''
        
        summary_start = subtotal_row + 2
        
        ws.cell(row=summary_start, column=1).value = 'AGING'
        ws.cell(row=summary_start, column=1).font = Font(bold=True)
        summary_start += 1
        
        # Only add Within 120 Days if non-zero
        if within_120 != 0:
            ws.cell(row=summary_start, column=1).value = 'Within 120 Days - PPW'
            cell_within = ws.cell(row=summary_start, column=2)
            cell_within.value = within_120
            cell_within.number_format = '#,##0.00'
            cell_within.font = Font(underline='single')
            summary_start += 1
        
        # Only add Over 120 Days if non-zero
        if over_120 != 0:
            ws.cell(row=summary_start, column=1).value = 'Over 120 Days'
            cell_over120 = ws.cell(row=summary_start, column=2)
            cell_over120.value = over_120
            cell_over120.number_format = '#,##0.00'
            cell_over120.font = Font(underline='single')
            summary_start += 1
        
        # Only add Over 180 Days if non-zero
        if over_180 != 0:
            ws.cell(row=summary_start, column=1).value = 'Over 180 Days'
            cell_over180 = ws.cell(row=summary_start, column=2)
            cell_over180.value = over_180
            cell_over180.number_format = '#,##0.00'
            cell_over180.font = Font(underline='single')
            summary_start += 1
        
        total_cell_1 = ws.cell(row=summary_start, column=1)
        total_cell_1.value = 'Total'
        total_cell_1.font = Font(bold=True)
        
        total_cell_2 = ws.cell(row=summary_start, column=2)
        total_cell_2.value = total_aging
        total_cell_2.font = Font(bold=True, underline='single')
        total_cell_2.number_format = '#,##0.00'
        
        footer_start_row = summary_start + 2
        italic_fmt = Font(italic=True)
        bold_fmt = Font(bold=True)
        regular_fmt = Font()
        
        italic_lines = {
            "For your convenience, payments may be made via the BDO Bills Payment facility:"
        }
        bold_lines = {
            "1. BDO Bills Payment", "a. BDO Mobile Application or BDO Web Page", "b. Over the Counter",
            "NOTE: Please make checks payable to PHILIPPINES FIRST INSURANCE CO., INC",
            "      Payments via LBP and BOC Peso are available only through special arrangement via fund transfer, with an advance copy of the remittance schedule."
        }
        footer_lines = [
            "For your convenience, payments may be made via the BDO Bills Payment facility:",
            "", "1. BDO Bills Payment", "a. BDO Mobile Application or BDO Web Page",
            "   i. Biller: Philippines First Insurance Co., Inc.", "   ii. Reference Number: HO-0001",
            "b. Over the Counter", "   i. Company Name: Philippines First Insurance Co., Inc.",
            "   ii. Subscriber Name: Your Company Name", "   iii. Subscriber Account Number: HO-0001",
            "", "NOTE: Please make checks payable to PHILIPPINES FIRST INSURANCE CO., INC",
            "      Payments via LBP and BOC Peso are available only through special arrangement via fund transfer, with an advance copy of the remittance schedule."
        ]
        for i, line in enumerate(footer_lines):
            cell = ws.cell(row=footer_start_row + i, column=1)
            cell.value = line
            if line in italic_lines:
                cell.font = italic_fmt
            elif line in bold_lines:
                cell.font = bold_fmt
            else:
                cell.font = regular_fmt
    
    for col_idx, cell in enumerate(ws[header_row], 1):
        header_val = cell.value
        col_letter = chr(64 + col_idx)
        
        if header_val in COLUMN_WIDTHS:
            ws.column_dimensions[col_letter].width = COLUMN_WIDTHS[header_val]
        else:
            ws.column_dimensions[col_letter].width = 15
    
    wb.save(file_path)

def load_merge_config():
    """Load merge configuration from merge_premium.py"""
    try:
        from soa_reinsurer.merge_reinsurers import merge_premium, premium_rename
        return merge_premium, premium_rename
    except (ImportError, AttributeError):
        return [], {}

def normalize_reinsurer_name(name):
    """Normalize reinsurer name for comparison"""
    return str(name).strip().upper()

def find_merge_group(reinsurer_name, merge_premium_list):
    """Find which merge group a reinsurer belongs to"""
    reinsurer_normalized = normalize_reinsurer_name(reinsurer_name)
    
    for group in merge_premium_list:
        for member in group:
            if normalize_reinsurer_name(member) == reinsurer_normalized:
                return group
    return None

def process_premium(files):
    """Process premium files with merge capability"""
    print("Processing Premium files...")
    
    df = pd.read_csv(files[0])
    df.columns = df.columns.str.strip()
    
    df['Aging'] = df.apply(determine_aging_premium, axis=1)
    df['REMARKS'] = ''
    
    current_date = datetime.now()
    month_short = current_date.strftime("%b").upper()
    year = current_date.strftime("%Y")
    updates_col_name = f'UPDATES_{month_short}. {year}'
    df[updates_col_name] = ''
    
    available_columns = [col for col in PREMIUM_COLUMNS if col in df.columns]
    if 'Aging' not in available_columns and 'Aging' in df.columns:
        available_columns.append('Aging')
    if 'REMARKS' not in available_columns and 'REMARKS' in df.columns:
        available_columns.append('REMARKS')
    available_columns.append(updates_col_name)
    
    df_processed = df[available_columns].copy()
    
    if 'Balance Due' in df_processed.columns:
        df_processed['Balance Due'] = pd.to_numeric(
            df_processed['Balance Due'].astype(str).str.replace(',', ''), 
            errors='coerce'
        )
    
    # === MODIFICATION 1: Remove rows with Balance Due = 0 (keep negatives) ===
    if 'Balance Due' in df_processed.columns:
        df_processed = df_processed[df_processed['Balance Due'] != 0].copy()
    
    merge_premium_list, premium_rename_map = load_merge_config()
    
    output_dfs = []
    processed_reinsurers = set()
    
    if 'Reinsurer' in df_processed.columns:
        reinsurers = df_processed['Reinsurer'].dropna().unique()
        
        for reinsurer in reinsurers:
            reinsurer_normalized = normalize_reinsurer_name(reinsurer)
            
            if reinsurer_normalized in processed_reinsurers:
                continue
            
            merge_group = find_merge_group(reinsurer, merge_premium_list)
            
            if merge_group:
                # Merge group processing - combine all members
                merged_sections = []
                master_name = None
                
                for group_member in merge_group:
                    group_member_data = df_processed[
                        df_processed['Reinsurer'].apply(
                            lambda x: normalize_reinsurer_name(x) == normalize_reinsurer_name(group_member)
                        )
                    ].copy()
                    
                    if not group_member_data.empty:
                        processed_reinsurers.add(normalize_reinsurer_name(group_member))
                        
                        if master_name is None:
                            master_name = premium_rename_map.get(group_member, group_member)
                        
                        # Get address for this group member
                        address = ''
                        if 'Address' in group_member_data.columns:
                            addr_val = group_member_data['Address'].iloc[0]
                            address = str(addr_val) if pd.notna(addr_val) else ''
                        
                        # === MODIFICATION 2: Check if subtotal = 0 before adding ===
                        if 'Balance Due' in group_member_data.columns:
                            subtotal = group_member_data['Balance Due'].sum()
                            if subtotal != 0:
                                merged_sections.append((group_member, group_member_data, address))
                        else:
                            merged_sections.append((group_member, group_member_data, address))
                
                if merged_sections and master_name:
                    output_dfs.append((master_name, merged_sections, True))
            else:
                # Single reinsurer processing
                reinsurer_df = df_processed[df_processed['Reinsurer'] == reinsurer].copy()
                
                # === MODIFICATION 2: Check if subtotal = 0 before adding ===
                if 'Balance Due' in reinsurer_df.columns:
                    subtotal = reinsurer_df['Balance Due'].sum()
                    if subtotal == 0:
                        print(f"  Skipping {reinsurer} - Balance Due subtotal is 0")
                        processed_reinsurers.add(reinsurer_normalized)
                        continue
                
                processed_reinsurers.add(reinsurer_normalized)
                
                address = ''
                if 'Address' in reinsurer_df.columns:
                    addr_val = reinsurer_df['Address'].iloc[0]
                    address = str(addr_val) if pd.notna(addr_val) else ''
                
                # Don't add total row - formatting function will calculate it
                output_dfs.append((reinsurer, reinsurer_df, False))
    
    return output_dfs

def extract_soa_reinsurer_premium(files):
    """
    Process SOA for premium reinsurer
    
    Args:
        files: List of uploaded files
    
    Returns:
        Tuple of (zip_path, zip_filename, temp_dir)
    """
    if not files or all(f.filename == "" for f in files):
        return None
    
    if len(files) < 1:
        raise ValueError("Premium requires at least 1 file")
    
    temp_dir = tempfile.mkdtemp()
    today = datetime.now().strftime("%b %d, %Y")  # e.g., "Jan 15, 2025"
    
    try:
        output_dfs = process_premium(files)
        excel_files = []
        
        for item in output_dfs:
            if len(item) == 3:
                filename, data, is_merged = item
                
                if is_merged:
                    # data is a list of (reinsurer_name, df, address) tuples
                    clean_name = make_filename_safe(filename)
                    file_name = f"SOA {clean_name} AS OF {today}.xlsx"
                    file_path = os.path.join(temp_dir, file_name)
                    
                    # Create initial workbook with merged sections
                    initial_df = pd.DataFrame()
                    initial_df.to_excel(file_path, index=False, engine='openpyxl')
                    
                    # Apply merged formatting with all sections
                    apply_premium_formatting_merged(file_path, data)
                    
                    excel_files.append(file_path)
                    print(f"✓ Created: {file_name} (Merged - {len(data)} reinsurers)")
                else:
                    # Single reinsurer
                    clean_name = make_filename_safe(filename)
                    file_name = f"SOA {clean_name} AS OF {today}.xlsx"
                    file_path = os.path.join(temp_dir, file_name)
                    
                    data.to_excel(file_path, index=False, engine='openpyxl')
                    apply_premium_formatting(file_path, filename, is_merged=False)
                    
                    excel_files.append(file_path)
                    print(f"✓ Created: {file_name}")
        
        zip_filename = f"SOA PREMIUM AS OF {today}.zip"
        zip_path = os.path.join(temp_dir, zip_filename)
        
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            for file_path in excel_files:
                zipf.write(file_path, os.path.basename(file_path))
        
        print(f"\n✓ ZIP file created: {zip_filename}")
        print(f"✓ Total files in ZIP: {len(excel_files)}")
        
        return zip_path, zip_filename, temp_dir
    
    except Exception as e:
        import shutil
        shutil.rmtree(temp_dir, ignore_errors=True)
        raise e