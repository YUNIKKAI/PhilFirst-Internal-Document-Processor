import os
import tempfile
import zipfile
import pandas as pd
from datetime import datetime
import re
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# Constants
PREMIUM_COLUMNS = [
    'Reinsurer', 'Address', 'Currency', 'Currency Rate', 'Line', 'Date', 'Our Policy No.', 'Invoice No.', 'Bord Date', 'Inst No.', 
    'Due Date', 'Assured', 'Assured Policy No.', 'Binder No.', 'Balance Due', 'Aging', 'REMARKS'
]

CASHCALL_COLUMNS = [
    'Branch', 'Line', 'Reinsurer', 'Assured', 'Policy Number', 
    'Claim Number', 'FLA Number', 'FLA Date', 'Loss Date', 'Aging',
    'Total Share', 'Total Payments', 'Total Amount Due'
]

def make_filename_safe(name: str) -> str:
    """Clean reinsurer name for use in filenames"""
    name = str(name).strip()
    name = re.sub(r'[<>:"/\\|?*]', '', name)
    name = re.sub(r'\s+', ' ', name)
    return name[:100]

def determine_aging_premium(row):
    """Determine aging based on which column has numerical data for Premium"""
    columns = ['CURRENT', 'OVER 30 DAYS', 'OVER 60 DAYS', 'OVER 90 DAYS', 'OVER 120 DAYS', 'OVER 180 DAYS']
    
    for col in columns:
        if col in row and pd.notna(row[col]) and str(row[col]).strip() != '-':
            try:
                val = float(str(row[col]).replace(',', ''))
                if val != 0:
                    if col in ['CURRENT', 'OVER 30 DAYS', 'OVER 60 DAYS', 'OVER 90 DAYS']:
                        return 'Within 120days-PPW'
                    elif col == 'OVER 120 DAYS':
                        return 'OVER 120 DAYS'
                    elif col == 'OVER 180 DAYS':
                        return 'OVER 180 DAYS'
            except:
                continue
    
    return 'Within 120days-PPW'

def calculate_aging_cashcall(fla_date):
    """Calculate aging based on FLA Date for cash call"""
    try:
        if pd.isna(fla_date):
            return 'CURRENT'
        
        if isinstance(fla_date, str):
            fla_date = pd.to_datetime(fla_date)
        
        today = datetime.now()
        days_diff = (today - fla_date).days
        
        if days_diff <= 30:
            return 'CURRENT'
        elif days_diff <= 60:
            return 'Over 30 days'
        elif days_diff <= 90:
            return 'Over 60 days'
        elif days_diff <= 120:
            return 'Over 90 days'
        elif days_diff <= 180:
            return 'Over 120 days'
        elif days_diff <= 360:
            return 'Over 180 days'
        else:
            return 'Over 360 days'
    except:
        return 'CURRENT'

def apply_premium_formatting(file_path, reinsurer_name):
    """Apply formatting to premium Excel file"""
    wb = load_workbook(file_path)
    ws = wb.active
    
    # Define styles
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    header_font = Font(bold=True)
    center_align = Alignment(horizontal='center', vertical='center')
    
    # Add company header (10 rows above data)
    ws.insert_rows(1, 10)
    
    # Row 1: Company Name
    ws['A1'] = 'PHILIPPINE FIRST INSURANCE CO. INC'
    ws['A1'].font = Font(bold=True, size=12)
    
    # Row 2: Statement of Account
    ws['A2'] = 'STATEMENT OF ACCOUNT'
    ws['A2'].font = Font(bold=True)
    
    # Row 3: Date
    today = datetime.now().strftime("AS OF %B %d, %Y").upper()
    ws['A3'] = today
    ws['A3'].font = Font(bold=True, color="0000FF")
    
    # Row 4: Empty
    
    # Row 5: New Facultative Premium
    ws['A5'] = 'NEW FACULTATIVE PREMIUM'
    ws['A5'].font = Font(bold=True)
    
    # Row 6: Empty
    
    # Row 7: Reinsurer
    ws['A7'] = reinsurer_name
    ws['A7'].font = Font(bold=True)
    
    # Row 8: Address - Get from first row of data
    address = ''
    address_col = None
    for idx, cell in enumerate(ws[11], start=1):
        if ws.cell(row=11, column=idx).value == 'Address':
            address_col = idx
            break
    
    if address_col and ws.max_row > 11:
        # Get address from first data row (row 12)
        address_value = ws.cell(row=12, column=address_col).value
        if address_value:
            address = str(address_value)
    
    ws['A8'] = address
    ws['A8'].font = Font(size=10)
    
    # Row 9: Empty
    
    # Row 10: Empty (reserved for future use)
    
    # Headers are now on row 11
    header_row = 11
    
    # Style headers
    for cell in ws[header_row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border
    
    # Find the subtotal row
    subtotal_row = None
    for idx, row in enumerate(ws.iter_rows(min_row=12, max_col=1), start=12):
        if row[0].value and 'TOTAL -' in str(row[0].value):
            subtotal_row = idx
            break
    
    if subtotal_row:
        # Apply border to all data rows
        max_col = ws.max_column
        for row in range(12, subtotal_row + 1):
            for col in range(1, max_col + 1):
                ws.cell(row=row, column=col).border = thin_border
        
        # Bold and highlight subtotal row
        for col in range(1, max_col + 1):
            cell = ws.cell(row=subtotal_row, column=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        
        # Calculate aging summary values
        aging_col = None
        balance_col = None
        
        for idx, cell in enumerate(ws[header_row], start=1):
            if cell.value == 'Aging':
                aging_col = idx
            elif cell.value == 'Balance Due':
                balance_col = idx
        
        within_120 = 0
        over_120 = 0
        over_180 = 0
        
        if aging_col and balance_col:
            for row in range(12, subtotal_row):
                aging_val = ws.cell(row=row, column=aging_col).value
                balance_val = ws.cell(row=row, column=balance_col).value
                
                if balance_val and isinstance(balance_val, (int, float)):
                    if aging_val == 'Within 120days-PPW':
                        within_120 += balance_val
                    elif aging_val == 'OVER 120 DAYS':
                        over_120 += balance_val
                    elif aging_val == 'OVER 180 DAYS':
                        over_180 += balance_val
        
        total_aging = within_120 + over_120 + over_180
        
        # Add aging summary (start 2 rows after subtotal)
        summary_start = subtotal_row + 2
        
        ws.cell(row=summary_start, column=1).value = 'AGING'
        ws.cell(row=summary_start, column=1).font = Font(bold=True)
        
        ws.cell(row=summary_start + 1, column=1).value = 'Within 120 Days - PPW'
        ws.cell(row=summary_start + 1, column=2).value = within_120
        ws.cell(row=summary_start + 1, column=2).number_format = '#,##0.00'
        
        ws.cell(row=summary_start + 2, column=1).value = 'Over 120 Days'
        ws.cell(row=summary_start + 2, column=2).value = over_120
        ws.cell(row=summary_start + 2, column=2).number_format = '#,##0.00'
        
        ws.cell(row=summary_start + 3, column=1).value = 'Over 180 Days'
        ws.cell(row=summary_start + 3, column=2).value = over_180
        ws.cell(row=summary_start + 3, column=2).number_format = '#,##0.00'
        
        ws.cell(row=summary_start + 4, column=1).value = 'Total'
        ws.cell(row=summary_start + 4, column=1).font = Font(bold=True)
        ws.cell(row=summary_start + 4, column=2).value = total_aging
        ws.cell(row=summary_start + 4, column=2).font = Font(bold=True)
        ws.cell(row=summary_start + 4, column=2).number_format = '#,##0.00'
        
        # Apply border to aging summary
        for row in range(summary_start, summary_start + 5):
            for col in range(1, 3):
                ws.cell(row=row, column=col).border = thin_border
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 40
    
    wb.save(file_path)

def process_premium(files):
    """Process premium files"""
    print("Processing Premium files...")
    
    df = pd.read_csv(files[0])
    
    df['Aging'] = df.apply(determine_aging_premium, axis=1)
    df['REMARKS'] = ''
    
    available_columns = [col for col in PREMIUM_COLUMNS if col in df.columns]
    df_processed = df[available_columns].copy()
    
    if 'Balance Due' in df_processed.columns:
        df_processed['Balance Due'] = pd.to_numeric(
            df_processed['Balance Due'].astype(str).str.replace(',', ''), 
            errors='coerce'
        )
    
    output_dfs = []
    if 'Reinsurer' in df_processed.columns:
        reinsuers = df_processed['Reinsurer'].dropna().unique()
        
        for reinsurer in reinsuers:
            reinsurer_df = df_processed[df_processed['Reinsurer'] == reinsurer].copy()
            
            if 'Balance Due' in reinsurer_df.columns:
                total_balance = reinsurer_df['Balance Due'].sum()
                
                total_row = pd.DataFrame({col: [''] for col in reinsurer_df.columns})
                total_row['Reinsurer'] = f'TOTAL - {reinsurer}'
                total_row['Balance Due'] = total_balance
                
                reinsurer_with_total = pd.concat([reinsurer_df, total_row], ignore_index=True)
                output_dfs.append((reinsurer, reinsurer_with_total))
            else:
                output_dfs.append((reinsurer, reinsurer_df))
    
    return output_dfs

def process_cashcall(files):
    """Process cashcall files (requires 2 files: bulk and summary)"""
    print("Processing Cash Call files...")
    
    bulk_file = None
    summary_file = None
    
    for file in files:
        filename_lower = file.filename.lower()
        if 'bulk' in filename_lower:
            bulk_file = file
        elif 'summary' in filename_lower:
            summary_file = file
    
    if not bulk_file or not summary_file:
        bulk_file = files[0]
        summary_file = files[1] if len(files) > 1 else None
    
    if not summary_file:
        raise ValueError("Cash Call requires 2 files: bulk data and summary with loss date")
    
    df_bulk = pd.read_csv(bulk_file)
    df_details = pd.read_csv(summary_file)
    
    df_bulk_copy = df_bulk.copy()
    df_details_copy = df_details.copy()
    
    df_bulk_copy['match_reinsurer'] = df_bulk_copy['Reinsurer'].astype(str).str.strip().str.upper()
    df_bulk_copy['match_policy'] = df_bulk_copy['Policy Number'].astype(str).str.strip().str.upper()
    df_bulk_copy['match_claim'] = df_bulk_copy['Claim Number'].astype(str).str.strip().str.upper()
    df_bulk_copy['match_fla_date'] = pd.to_datetime(df_bulk_copy['FLA Date'], errors='coerce')
    
    df_details_copy['match_reinsurer'] = df_details_copy['REINSURER'].astype(str).str.strip().str.upper()
    df_details_copy['match_policy'] = df_details_copy['ASSURED POLICY NO'].astype(str).str.strip().str.upper()
    df_details_copy['match_claim'] = df_details_copy['CLAIM NO'].astype(str).str.strip().str.upper()
    df_details_copy['match_fla_date'] = pd.to_datetime(df_details_copy['FLA DATE'], errors='coerce')
    
    merged_df = df_bulk_copy.merge(
        df_details_copy[['match_reinsurer', 'match_policy', 'match_claim', 'match_fla_date', 'LOSS DATE']],
        on=['match_reinsurer', 'match_policy', 'match_claim', 'match_fla_date'],
        how='left'
    )
    
    df_bulk_copy['Loss Date'] = merged_df['LOSS DATE']
    df_bulk_copy['Aging'] = df_bulk_copy['FLA Date'].apply(calculate_aging_cashcall)
    
    available_columns = [col for col in CASHCALL_COLUMNS if col in df_bulk_copy.columns]
    df_processed = df_bulk_copy[available_columns].copy()
    
    if 'Total Amount Due' in df_processed.columns:
        df_processed['Total Amount Due'] = pd.to_numeric(
            df_processed['Total Amount Due'].astype(str).str.replace(',', ''), 
            errors='coerce'
        )
    
    output_dfs = []
    if 'Reinsurer' in df_processed.columns:
        reinsuers = df_processed['Reinsurer'].dropna().unique()
        
        for reinsurer in reinsuers:
            reinsurer_df = df_processed[df_processed['Reinsurer'] == reinsurer].copy()
            
            if 'Total Amount Due' in reinsurer_df.columns:
                total_amount = reinsurer_df['Total Amount Due'].sum()
                
                total_row = pd.DataFrame({col: [''] for col in reinsurer_df.columns})
                total_row['Reinsurer'] = f'TOTAL - {reinsurer}'
                total_row['Total Amount Due'] = total_amount
                
                reinsurer_with_total = pd.concat([reinsurer_df, total_row], ignore_index=True)
                output_dfs.append((reinsurer, reinsurer_with_total))
            else:
                output_dfs.append((reinsurer, reinsurer_df))
    
    return output_dfs

def extract_soa_reinsurer(files, file_type=None):
    """
    Main function to process SOA for reinsurer
    
    Args:
        files: List of uploaded files
        file_type: 'premium' or 'cash-call'
    
    Returns:
        Tuple of (zip_path, zip_filename, temp_dir)
    """
    if not files or all(f.filename == "" for f in files):
        return None
    
    if not file_type:
        return None
    
    temp_dir = tempfile.mkdtemp()
    today = datetime.now().strftime("%m-%d-%Y")
    
    try:
        if file_type == 'premium':
            if len(files) < 1:
                raise ValueError("Premium requires at least 1 file")
            output_dfs = process_premium(files)
        elif file_type == 'cash-call':
            if len(files) < 2:
                raise ValueError("Cash Call requires 2 files (bulk and summary)")
            output_dfs = process_cashcall(files)
        else:
            raise ValueError(f"Invalid file type: {file_type}")
        
        excel_files = []
        for reinsurer_name, reinsurer_df in output_dfs:
            if reinsurer_df.empty:
                continue
            
            clean_name = make_filename_safe(reinsurer_name)
            
            file_name = f"SoA of {clean_name} as of {today}.xlsx"
            file_path = os.path.join(temp_dir, file_name)
            
            reinsurer_df.to_excel(file_path, index=False, engine='openpyxl')
            
            # Apply formatting for premium files
            if file_type == 'premium':
                apply_premium_formatting(file_path, reinsurer_name)
            
            excel_files.append(file_path)
            print(f"✓ Created: {file_name}")
        
        zip_filename = f"SoA Reinsurance as of {today}.zip"
        zip_path = os.path.join(temp_dir, zip_filename)
        
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            for file_path in excel_files:
                zipf.write(file_path, os.path.basename(file_path))
        
        print(f"\n✓ ZIP file created: {zip_filename}")
        print(f"✓ Total files in ZIP: {len(excel_files)}")
        
        return zip_path, zip_filename, temp_dir
    
    except Exception as e:
        import shutil
        shutil.rmtree(temp_dir, ignore_errors=True)
        raise e