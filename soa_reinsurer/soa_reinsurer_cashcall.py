import os
import tempfile
import zipfile
import pandas as pd
from datetime import datetime
import re
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side

# Constants
CASHCALL_COLUMNS = [ 
    'Branch', 'Line', 'Reinsurer', 'Assured', 'Policy Number',
    'Claim Number', 'FLA Number', 'FLA Date', 'Loss Date', 'Aging',
    'Total Amount Due'
]

COLUMN_WIDTHS = {
    'Branch': 13,
    'Line': 11,
    'Reinsurer': 17.875,
    'Assured': 21.5,
    'Policy Number': 15,
    'Claim Number': 12,
    'FLA Number': 12,
    'FLA Date': 11,
    'Loss Date': 11,
    'Aging': 13.25,
    'Total Amount Due': 13
}

def make_filename_safe(name: str) -> str:
    """Clean reinsurer name for use in filenames"""
    name = str(name).strip()
    name = re.sub(r'[<>:"/\\|?*]', '', name)
    name = re.sub(r'\s+', ' ', name)
    return name[:100]

def calculate_aging_cashcall(fla_date):
    """Calculate aging based on FLA Date for cash call"""
    try:
        if pd.isna(fla_date):
            return 'CURRENT'
        
        if isinstance(fla_date, str):
            fla_date = pd.to_datetime(fla_date, format='mixed', dayfirst=False)
        
        today = datetime.now()
        days_diff = (today - fla_date).days
        
        if days_diff <= 30:
            return 'CURRENT'
        elif days_diff <= 60:
            return 'Over 30 days'
        elif days_diff <= 90:
            return 'Over 60 days'
        elif days_diff <= 120:
            return 'Over 90 days'
        elif days_diff <= 180:
            return 'Over 120 days'
        elif days_diff <= 360:
            return 'Over 180 days'
        else:
            return 'Over 360 days'
    except:
        return 'CURRENT'

def clean_cashcall_bulk_data(df):
    """Remove rows where only 'Assured' has data and handle partial Policy Number rows"""
    print("Cleaning bulk data...")
    original_count = len(df)

    # Defensive copy
    df = df.copy()

    # Ensure Policy Number column exists
    if 'Policy Number' not in df.columns:
        raise ValueError("'Policy Number' column is missing in bulk data")

    # === Step 1: Handle policy numbers ending with '-' ===
    print("Processing incomplete policy numbers (ending with '-')...")
    rows_to_drop_step1 = []
    
    for idx in df.index:
        row = df.loc[idx]
        policy_num = str(row.get('Policy Number', '')).strip()
        
        # Check if policy number ends with '-'
        if policy_num.endswith('-'):
            print(f"  Found incomplete policy at row {idx}: {policy_num}")
            
            # Look ahead to find the next row with policy number
            current_index_pos = df.index.get_loc(idx)
            next_policy = None
            next_idx = None
            
            # Search following rows for a row with only policy number
            for search_offset in range(1, len(df) - current_index_pos):
                search_idx = df.index[current_index_pos + search_offset]
                search_row = df.loc[search_idx]
                
                search_policy = str(search_row.get('Policy Number', '')).strip()
                has_reinsurer = not pd.isna(search_row.get('Reinsurer')) and str(search_row.get('Reinsurer')).strip() != ''
                has_claim = not pd.isna(search_row.get('Claim Number')) and str(search_row.get('Claim Number')).strip() != ''
                
                # If we hit another complete row, stop searching
                if has_reinsurer or has_claim:
                    break
                
                # Found a row with policy number (likely the continuation)
                if search_policy and search_policy != 'nan':
                    next_policy = search_policy
                    next_idx = search_idx
                    print(f"    Found continuation at row {search_idx}: {next_policy}")
                    break
            
            # Concatenate the next policy number to the incomplete one
            if next_policy and next_idx is not None:
                # Keep the trailing '-' and concatenate directly (no space)
                new_policy = policy_num + next_policy
                df.at[idx, 'Policy Number'] = new_policy.strip()
                print(f"    Concatenated: {policy_num} + {next_policy} = {new_policy}")
                
                # Mark the continuation row for deletion
                rows_to_drop_step1.append(next_idx)
                print(f"    Marked row {next_idx} for removal")
            else:
                print(f"    No continuation found for incomplete policy at row {idx}")
    
    # Drop concatenated continuation rows
    if rows_to_drop_step1:
        print(f"  Removing {len(rows_to_drop_step1)} continuation rows...")
        df = df.drop(rows_to_drop_step1).reset_index(drop=True)

    # === Step 2: Handle rows with only Assured and/or Policy Number ===
    rows_to_drop = []
    last_valid_index = None  # Remember last "complete" row

    for idx in df.index:
        row = df.loc[idx]

        has_reinsurer = not pd.isna(row.get('Reinsurer')) and str(row.get('Reinsurer')).strip() != ''
        has_claim = not pd.isna(row.get('Claim Number')) and str(row.get('Claim Number')).strip() != ''
        has_policy = not pd.isna(row.get('Policy Number')) and str(row.get('Policy Number')).strip() != ''
        has_assured = not pd.isna(row.get('Assured')) and str(row.get('Assured')).strip() != ''

        # Complete row — update pointer
        if has_reinsurer or has_claim:
            last_valid_index = idx
            continue

        # Only Assured → drop
        if has_assured and not has_policy:
            rows_to_drop.append(idx)
            continue

        # Has Policy Number (and maybe Assured) → concatenate to last complete row
        if has_policy and last_valid_index is not None:
            prev_policy = str(df.at[last_valid_index, 'Policy Number'])
            new_policy = str(row['Policy Number']).strip()
            if new_policy not in prev_policy:
                df.at[last_valid_index, 'Policy Number'] = f"{prev_policy} {new_policy}".strip()
            rows_to_drop.append(idx)
            continue

        # Neither assured nor policy number → drop
        if not has_assured and not has_policy:
            rows_to_drop.append(idx)

    # Drop those filler rows
    df = df.drop(rows_to_drop).reset_index(drop=True)

    # === Step 3: Continue original logic ===
    critical_columns = ['Reinsurer', 'Policy Number', 'Claim Number']
    mask = pd.Series([True] * len(df))

    for col in critical_columns:
        if col in df.columns:
            col_empty = df[col].isna() | (df[col].astype(str).str.strip() == '') | (df[col].astype(str).str.strip() == 'nan')
            mask = mask & col_empty

    df_cleaned = df[~mask].copy()

    removed_count = original_count - len(df_cleaned)
    print(f"  Original rows: {original_count}")
    print(f"  Removed rows: {removed_count}")
    print(f"  Remaining rows: {len(df_cleaned)}")

    return df_cleaned

def load_merge_config():
    """Load merge configuration from merge_cashcall.py"""
    try:
        from soa_reinsurer.merge_reinsurers import merge_cashcall, cashcall_rename
        return merge_cashcall, cashcall_rename
    except (ImportError, AttributeError):
        return [], {}

def normalize_reinsurer_name(name):
    """Normalize reinsurer name for comparison"""
    return str(name).strip().upper()

def find_merge_group(reinsurer_name, merge_list):
    """Find which merge group a reinsurer belongs to"""
    reinsurer_normalized = normalize_reinsurer_name(reinsurer_name)
    
    for group in merge_list:
        for member in group:
            if normalize_reinsurer_name(member) == reinsurer_normalized:
                return group
    return None

def process_cashcall(files):
    """Process cashcall files (requires 2 files: bulk and summary)"""
    print("Processing Cash Call files...")
    
    bulk_file = None
    summary_file = None
    
    for file in files:
        filename_lower = file.filename.lower()
        print(f"Checking file: {file.filename}")
        if 'bulk' in filename_lower:
            bulk_file = file
            print(f"  -> Identified as BULK")
        elif 'summary' in filename_lower:
            summary_file = file
            print(f"  -> Identified as SUMMARY")
    
    if not bulk_file or not summary_file:
        print("Auto-assigning files based on position...")
        bulk_file = files[0]
        summary_file = files[1] if len(files) > 1 else None
        print(f"  Bulk: {bulk_file.filename}")
        print(f"  Summary: {summary_file.filename if summary_file else 'None'}")
    
    if not summary_file:
        raise ValueError("Cash Call requires 2 files: bulk data and summary with loss date")
    
    print("Reading CSV files...")
    df_bulk = pd.read_csv(bulk_file)
    df_bulk.columns = df_bulk.columns.str.strip()
    df_details = pd.read_csv(summary_file, header=7)
    df_details.columns = df_details.columns.str.strip()
    print(f"Bulk file rows (before cleaning): {len(df_bulk)}")
    print(f"Summary file rows: {len(df_details)}")
    print(f"Bulk columns: {list(df_bulk.columns)}")
    
    df_bulk = clean_cashcall_bulk_data(df_bulk)
    
    if len(df_bulk) == 0:
        raise ValueError("No valid data found in bulk file after cleaning")
    
    df_bulk_copy = df_bulk.copy()
    df_details_copy = df_details.copy()
    
    print("Parsing FLA dates...")
    df_bulk_copy['match_fla_date'] = pd.to_datetime(
        df_bulk_copy['FLA Date'], 
        format='mixed',
        dayfirst=False,
        errors='coerce'
    )
    
    df_details_copy['match_fla_date'] = pd.to_datetime(
        df_details_copy['FLA DATE'], 
        format='mixed',
        dayfirst=False,
        errors='coerce'
    )
    
    print("Merging dataframes on FLA Date match only...")
    merged_df = df_bulk_copy.merge(
        df_details_copy[['match_fla_date', 'LOSS DATE']],
        on=['match_fla_date'],
        how='left'
    )
    print(f"Merged rows: {len(merged_df)}")
    
    print("Populating Loss Date with matched values or '-'...")
    df_bulk_copy['Loss Date'] = merged_df['LOSS DATE'].fillna('-')
    print(f"  Loss Date populated: {len(df_bulk_copy)} rows")
    print(f"  Rows with Loss Date value: {(df_bulk_copy['Loss Date'] != '-').sum()}")
    print(f"  Rows with '-': {(df_bulk_copy['Loss Date'] == '-').sum()}")
    
    print("Calculating aging...")
    df_bulk_copy['Aging'] = df_bulk_copy['FLA Date'].apply(calculate_aging_cashcall)
    
    print(f"Columns in df_bulk_copy: {list(df_bulk_copy.columns)}")
    
    available_columns = [col for col in CASHCALL_COLUMNS if col in df_bulk_copy.columns]
    print(f"CASHCALL_COLUMNS: {CASHCALL_COLUMNS}")
    print(f"Actual columns in df_bulk_copy: {list(df_bulk_copy.columns)}")
    print(f"Available columns after filter: {available_columns}")
    df_processed = df_bulk_copy[available_columns].copy()
    
    # CRITICAL: Process Total Amount Due - preserve negatives!
    if 'Total Amount Due' in df_processed.columns:
        print("Processing Total Amount Due - PRESERVING NEGATIVES...")
        print(f"  Before conversion - sample raw values: {df_processed['Total Amount Due'].head(10).tolist()}")

        # Convert to a string series for manipulation
        amount_series = df_processed['Total Amount Due'].astype(str).str.strip()

        # 1. Remove commas
        amount_series = amount_series.str.replace(',', '', regex=False)
        # 2. Replace parentheses with a standard negative sign
        amount_series = amount_series.str.replace('(', '-', regex=False).str.replace(')', '', regex=False)
        # 3. Handle trailing minus signs (e.g., "123.45-")
        amount_series = amount_series.apply(lambda x: f"-{x[:-1]}" if x.endswith('-') else x)

        # Now, convert to numeric. With the cleaning above, this will work correctly.
        df_processed['Total Amount Due'] = pd.to_numeric(amount_series, errors='coerce')

        print(f"  After conversion - sample values: {df_processed['Total Amount Due'].head(10).tolist()}")
        print(f"  Negative count: {(df_processed['Total Amount Due'] < 0).sum()}")
        print(f"  Positive count: {(df_processed['Total Amount Due'] > 0).sum()}")
        print(f"  Zero count: {(df_processed['Total Amount Due'] == 0).sum()}")
        
    merge_list, rename_map = load_merge_config()
    
    output_dfs = []
    processed_reinsurers = set()
    
    if 'Reinsurer' in df_processed.columns:
        reinsurers = df_processed['Reinsurer'].dropna().unique()
        
        for reinsurer in reinsurers:
            reinsurer_normalized = normalize_reinsurer_name(reinsurer)
            
            if reinsurer_normalized in processed_reinsurers:
                continue
            
            merge_group = find_merge_group(reinsurer, merge_list)
            
            if merge_group:
                # Merge group: combine all members
                merged_sections = []
                master_name = None
                total_for_merged = 0
                
                for group_member in merge_group:
                    group_member_data = df_processed[
                        df_processed['Reinsurer'].apply(
                            lambda x: normalize_reinsurer_name(x) == normalize_reinsurer_name(group_member)
                        )
                    ].copy()
                    
                    if not group_member_data.empty:
                        processed_reinsurers.add(normalize_reinsurer_name(group_member))
                        
                        if master_name is None:
                            master_name = rename_map.get(group_member, group_member)
                        
                        # Calculate subtotal for this member (INCLUDING NEGATIVES)
                        if 'Total Amount Due' in group_member_data.columns:
                            member_total = group_member_data['Total Amount Due'].sum()
                            total_for_merged += member_total
                            print(f"  Group member {group_member}: total = {member_total}")
                        
                        merged_sections.append((group_member, group_member_data))
                
                # Only add if total is not 0
                if merged_sections and master_name and total_for_merged != 0:
                    # Now remove rows with 0 amount from each section
                    filtered_sections = []
                    for section_name, section_df in merged_sections:
                        if 'Total Amount Due' in section_df.columns:
                            section_df_filtered = section_df[section_df['Total Amount Due'] != 0].copy()
                            if not section_df_filtered.empty:
                                filtered_sections.append((section_name, section_df_filtered))
                        else:
                            filtered_sections.append((section_name, section_df))
                    
                    if filtered_sections:
                        output_dfs.append((master_name, filtered_sections, True))
                        print(f"  Added merged group: {master_name} (Total: {total_for_merged})")
                elif total_for_merged == 0:
                    print(f"  Skipped merged group: {master_name} (Total Amount Due = 0)")
            else:
                # Single reinsurer
                reinsurer_df = df_processed[df_processed['Reinsurer'] == reinsurer].copy()
                processed_reinsurers.add(reinsurer_normalized)
                
                # Check if subtotal is 0
                if 'Total Amount Due' in reinsurer_df.columns:
                    subtotal = reinsurer_df['Total Amount Due'].sum()
                    print(f"  Reinsurer {reinsurer}: subtotal = {subtotal}")
                    if subtotal != 0:
                        # Remove rows with 0 amount
                        reinsurer_df_filtered = reinsurer_df[reinsurer_df['Total Amount Due'] != 0].copy()
                        if not reinsurer_df_filtered.empty:
                            output_dfs.append((reinsurer, reinsurer_df_filtered, False))
                            print(f"  Added single reinsurer: {reinsurer} (Total: {subtotal})")
                    else:
                        print(f"  Skipped reinsurer: {reinsurer} (Total Amount Due = 0)")
                else:
                    output_dfs.append((reinsurer, reinsurer_df, False))
    
    print(f"Created {len(output_dfs)} output dataframes")
    return output_dfs

def apply_cashcall_formatting(file_path, reinsurer_name):
    """Apply formatting to single cashcall Excel file"""
    print(f"\n=== FORMATTING SINGLE FILE: {reinsurer_name} ===")
    wb = load_workbook(file_path)
    ws = wb.active
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    header_font = Font(bold=True)
    center_align = Alignment(horizontal='center', vertical='center')
    
    # Insert header rows - header starts at row 9
    ws.insert_rows(1, 8)
    header_row = 9
    
    # Add titles
    ws['A1'] = 'PHILIPPINE FIRST INSURANCE CO. INC'
    ws['A1'].font = Font(bold=True, size=12)
    
    ws['A2'] = 'STATEMENT OF ACCOUNT'
    ws['A2'].font = Font(bold=True)
    
    today = datetime.now().strftime("AS OF %B %d, %Y").upper()
    ws['A3'] = today
    ws['A3'].font = Font(bold=True)
    
    ws['A5'] = 'NEW CASH CALL'
    ws['A5'].font = Font(bold=True)
    
    ws['A7'] = reinsurer_name
    ws['A7'].font = Font(bold=True)
    
    # Format header row
    for col_idx, cell in enumerate(ws[header_row], 1):
        if cell.value:
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border
    
    # Apply borders to data rows
    data_start_row = header_row + 1
    data_end_row = ws.max_row
    
    for row in range(data_start_row, data_end_row + 1):
        for col in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col).border = thin_border
    
    # Set column widths
    for col_idx, cell in enumerate(ws[header_row], 1):
        header_val = cell.value
        col_letter = chr(64 + col_idx)
        
        if header_val in COLUMN_WIDTHS:
            ws.column_dimensions[col_letter].width = COLUMN_WIDTHS[header_val]
        else:
            ws.column_dimensions[col_letter].width = 15
    
    # Find Total Amount Due column index
    amount_col = None
    for col_idx, cell in enumerate(ws[header_row], 1):
        if cell.value == 'Total Amount Due':
            amount_col = col_idx
            break
    
    print(f"Amount column index: {amount_col}")
    print(f"Data range: rows {data_start_row} to {data_end_row}")
    
    # === PHASE 1: CALCULATION - NO FORMATTING ===
    print("\n--- PHASE 1: CALCULATIONS ---")
    total_amount = 0
    aging_summary = {
        'CURRENT': 0,
        'Over 30 days': 0,
        'Over 60 days': 0,
        'Over 90 days': 0,
        'Over 120 days': 0,
        'Over 180 days': 0,
        'Over 360 days': 0
    }
    
    aging_col = None
    for col_idx, cell in enumerate(ws[header_row], 1):
        if cell.value == 'Aging':
            aging_col = col_idx
            break
    
    # Calculate totals WITHOUT any formatting
    for row in range(data_start_row, data_end_row + 1):
        if amount_col:
            amount_cell = ws.cell(row=row, column=amount_col)
            cell_val = amount_cell.value
            
            if cell_val is not None and cell_val != '':
                if isinstance(cell_val, (int, float)):
                    print(f"  Row {row}: value = {cell_val} (type: {type(cell_val)})")
                    total_amount += cell_val
                else:
                    print(f"  Row {row}: value = '{cell_val}' (type: {type(cell_val)}) - SKIPPED")
        
        if aging_col and amount_col:
            aging_cell = ws.cell(row=row, column=aging_col)
            aging_val = aging_cell.value
            amount_cell = ws.cell(row=row, column=amount_col)
            
            if aging_val and amount_cell.value is not None:
                aging_str = str(aging_val).strip()
                if aging_str in aging_summary:
                    try:
                        amt = float(amount_cell.value) if isinstance(amount_cell.value, (int, float)) else float(str(amount_cell.value).replace(',', ''))
                        aging_summary[aging_str] += amt
                        print(f"  Aging {aging_str}: added {amt}")
                    except (ValueError, TypeError) as e:
                        print(f"  Aging calculation error: {e}")
    
    print(f"\nTotal calculated: {total_amount}")
    print(f"Aging summary: {aging_summary}")
    
    # Add subtotal row
    subtotal_row = data_end_row + 1
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=subtotal_row, column=col)
        cell.border = thin_border
        if col == amount_col:
            cell.value = total_amount
            cell.font = Font(bold=True)
    
    current_row = subtotal_row + 2
    
    # Add aging summary (only non-zero categories)
    ws.cell(row=current_row, column=1).value = 'AGING SUMMARY'
    ws.cell(row=current_row, column=1).font = Font(bold=True)
    current_row += 1
    
    for aging_label in ['CURRENT', 'Over 30 days', 'Over 60 days', 'Over 90 days', 'Over 120 days', 'Over 180 days', 'Over 360 days']:
        aging_value = aging_summary.get(aging_label, 0)
        if aging_value != 0:
            ws.cell(row=current_row, column=1).value = aging_label
            aging_cell = ws.cell(row=current_row, column=2)
            aging_cell.value = aging_value
            aging_cell.font = Font(underline='single')
            current_row += 1
    
    # Total aging row
    total_aging = sum(aging_summary.values())
    ws.cell(row=current_row, column=1).value = 'Total'
    ws.cell(row=current_row, column=1).font = Font(bold=True)
    total_aging_cell = ws.cell(row=current_row, column=2)
    total_aging_cell.value = total_aging
    total_aging_cell.font = Font(bold=True, underline='single')
    current_row += 2
    
    # === Add footer ===
    current_row += 1
    italic_fmt = Font(italic=True)
    bold_fmt = Font(bold=True)
    regular_fmt = Font()
    italic_lines = {
        "For your convenience, payments may be made via the BDO Bills Payment facility:"
    }
    bold_lines = {
        "1. BDO Bills Payment", "a. BDO Mobile Application or BDO Web Page", "b. Over the Counter",
        "NOTE: Please make checks payable to PHILIPPINES FIRST INSURANCE CO., INC",
        "      Payments via LBP and BOC Peso are available only through special arrangement via fund transfer, with an advance copy of the remittance schedule."
    }
    footer_lines = [
        "For your convenience, payments may be made via the BDO Bills Payment facility:",
        "", "1. BDO Bills Payment", "a. BDO Mobile Application or BDO Web Page",
        "   i. Biller: Philippines First Insurance Co., Inc.", "   ii. Reference Number: HO-0001",
        "b. Over the Counter", "   i. Company Name: Philippines First Insurance Co., Inc.",
        "   ii. Subscriber Name: Your Company Name", "   iii. Subscriber Account Number: HO-0001",
        "", "NOTE: Please make checks payable to PHILIPPINES FIRST INSURANCE CO., INC",
        "      Payments via LBP and BOC Peso are available only through special arrangement via fund transfer, with an advance copy of the remittance schedule."
    ]
    for line in footer_lines:
        cell = ws.cell(row=current_row, column=1)
        cell.value = line
        if line in italic_lines:
            cell.font = italic_fmt
        elif line in bold_lines:
            cell.font = bold_fmt
        else:
            cell.font = regular_fmt
        current_row += 1
    
    # === PHASE 2: FORMATTING ONLY ===
    print("\n--- PHASE 2: FORMATTING ---")
    
    # Define the standard accounting number format
    # Positive: #,##0.00; Negative: [Red](#,##0.00); Zero: 0.00
    accounting_format = '#,##0.00;[Red](#,##0.00);0.00'
    
    # Format data rows
    for row in range(data_start_row, data_end_row + 1):
        if amount_col:
            amount_cell = ws.cell(row=row, column=amount_col)
            if amount_cell.value is not None and isinstance(amount_cell.value, (int, float)):
                amount_cell.number_format = accounting_format
    
    # Format subtotal
    subtotal_cell = ws.cell(row=subtotal_row, column=amount_col)
    if subtotal_cell.value is not None and isinstance(subtotal_cell.value, (int, float)):
        subtotal_cell.number_format = accounting_format
        subtotal_cell.font = Font(bold=True) # Keep font bold, color is handled by format
    
    # Format aging summary values
    aging_summary_start = subtotal_row + 3
    # Note: current_row is now the row *after* the footer was added. We need the end of the aging block.
    aging_end_row = subtotal_row + 3 + len([v for v in aging_summary.values() if v != 0]) + 1

    for row in range(aging_summary_start, aging_end_row + 1):
        label_cell = ws.cell(row=row, column=1)
        value_cell = ws.cell(row=row, column=2)
        
        if value_cell.value is not None and isinstance(value_cell.value, (int, float)):
            value_cell.number_format = accounting_format
            
            # Apply underline/bold styling, color is handled by the format string
            if label_cell.value == 'Total':
                value_cell.font = Font(bold=True, underline='single')
            else:
                value_cell.font = Font(underline='single')
    
    print(f"=== FORMATTING COMPLETE ===\n")
    wb.save(file_path)

def apply_cashcall_formatting_merged(file_path, reinsurer_groups):
    """Apply formatting to merged cashcall Excel file with separate sections per reinsurer"""
    print(f"\n=== FORMATTING MERGED FILE ===")
    wb = load_workbook(file_path)
    ws = wb.active
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    header_font = Font(bold=True)
    center_align = Alignment(horizontal='center', vertical='center')
    separator_font = Font(bold=True, size=11)
    
    # Clear existing content
    ws.delete_rows(1, ws.max_row)
    
    current_row = 1
    
    # Header information
    ws.cell(row=current_row, column=1).value = 'PHILIPPINE FIRST INSURANCE CO. INC'
    ws.cell(row=current_row, column=1).font = Font(bold=True, size=12)
    current_row += 1
    
    ws.cell(row=current_row, column=1).value = 'STATEMENT OF ACCOUNT'
    ws.cell(row=current_row, column=1).font = Font(bold=True)
    current_row += 1
    
    today = datetime.now().strftime("AS OF %B %d, %Y").upper()
    ws.cell(row=current_row, column=1).value = today
    ws.cell(row=current_row, column=1).font = Font(bold=True)
    current_row += 1
    
    current_row += 1
    
    ws.cell(row=current_row, column=1).value = 'NEW CASH CALL'
    ws.cell(row=current_row, column=1).font = Font(bold=True)
    current_row += 1
    
    current_row += 1
    
    # Track grand totals for all sections
    grand_total_amount = 0
    
    # Store section positions for later formatting
    section_positions = []
    
    # === PHASE 1: DATA WRITING - NO FORMATTING ===
    print("\n--- PHASE 1: WRITING DATA ---")
    
    # Process each reinsurer section
    for group_idx, (reinsurer_name, reinsurer_df) in enumerate(reinsurer_groups):
        print(f"\nProcessing section: {reinsurer_name}")
        
        if group_idx > 0:
            # Separator between sections
            current_row += 1
            separator_cell = ws.cell(row=current_row, column=1)
            separator_cell.value = '.' * 100
            separator_cell.font = separator_font
            current_row += 2
        
        section_start_row = current_row
        
        # Reinsurer name
        ws.cell(row=current_row, column=1).value = reinsurer_name
        ws.cell(row=current_row, column=1).font = Font(bold=True)
        current_row += 1
        
        current_row += 1
        
        # Header row for this reinsurer's data
        header_row = current_row
        
        for col_idx, col_name in enumerate(reinsurer_df.columns, 1):
            cell = ws.cell(row=header_row, column=col_idx)
            cell.value = col_name
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border
            
            col_letter = chr(64 + col_idx)
            if col_name in COLUMN_WIDTHS:
                ws.column_dimensions[col_letter].width = COLUMN_WIDTHS[col_name]
            else:
                ws.column_dimensions[col_letter].width = 15
        
        current_row += 1
        data_start_row = current_row
        
        # Write data rows
        max_col = len(reinsurer_df.columns)
        total_amount = 0
        amount_col = None
        aging_col = None
        
        # Find column indices
        for idx, col_name in enumerate(reinsurer_df.columns, 1):
            if col_name == 'Total Amount Due':
                amount_col = idx
            elif col_name == 'Aging':
                aging_col = idx
        
        aging_summary = {
            'CURRENT': 0,
            'Over 30 days': 0,
            'Over 60 days': 0,
            'Over 90 days': 0,
            'Over 120 days': 0,
            'Over 180 days': 0,
            'Over 360 days': 0
        }
        
        # Write data rows - PRESERVE ALL VALUES
        for _, row_data in reinsurer_df.iterrows():
            for col_idx in range(1, max_col + 1):
                cell = ws.cell(row=current_row, column=col_idx)
                col_name = reinsurer_df.columns[col_idx - 1]
                value = row_data.iloc[col_idx - 1]
                
                # Write value as-is, NO FORMATTING
                if pd.notna(value) and value != '':
                    cell.value = value
                    
                    if col_name == 'Total Amount Due':
                        print(f"  Row {current_row}: Writing amount = {value} (type: {type(value)})")
                
                cell.border = thin_border
                
                # Track totals (including negatives)
                if col_name == 'Total Amount Due' and pd.notna(value) and value != '':
                    try:
                        num_val = float(str(value).replace(',', ''))
                        total_amount += num_val
                        grand_total_amount += num_val
                    except (ValueError, TypeError):
                        pass
                
                # Track aging
                if col_name == 'Aging' and aging_col and amount_col:
                    aging_val = str(value).strip() if pd.notna(value) else ''
                    amount_val = row_data.iloc[amount_col - 1]
                    
                    if aging_val in aging_summary and pd.notna(amount_val):
                        try:
                            amt = float(str(amount_val).replace(',', ''))
                            aging_summary[aging_val] += amt
                        except (ValueError, TypeError):
                            pass
            
            current_row += 1
        
        data_end_row = current_row - 1
        
        print(f"  Section total: {total_amount}")
        print(f"  Section aging summary: {aging_summary}")
        
        # Subtotal row
        subtotal_row = current_row
        for col in range(1, max_col + 1):
            cell = ws.cell(row=subtotal_row, column=col)
            cell.border = thin_border
            if col == amount_col:
                cell.value = total_amount
                cell.font = Font(bold=True)
            else:
                cell.value = ''
        
        current_row += 2
        
        # Aging summary for this section (only non-zero categories)
        aging_summary_start = current_row
        ws.cell(row=current_row, column=1).value = 'AGING SUMMARY'
        ws.cell(row=current_row, column=1).font = Font(bold=True)
        current_row += 1
        
        for aging_label in ['CURRENT', 'Over 30 days', 'Over 60 days', 'Over 90 days', 'Over 120 days', 'Over 180 days', 'Over 360 days']:
            aging_value = aging_summary.get(aging_label, 0)
            if aging_value != 0:
                ws.cell(row=current_row, column=1).value = aging_label
                aging_cell = ws.cell(row=current_row, column=2)
                aging_cell.value = aging_value
                aging_cell.font = Font(underline='single')
                current_row += 1
        
        # Total aging row for section
        total_section_aging = sum(aging_summary.values())
        total_aging_row = current_row
        ws.cell(row=current_row, column=1).value = 'Total'
        ws.cell(row=current_row, column=1).font = Font(bold=True)
        total_aging_cell = ws.cell(row=current_row, column=2)
        total_aging_cell.value = total_section_aging
        total_aging_cell.font = Font(bold=True, underline='single')
        current_row += 2
        
        # Store section info for formatting later
        section_positions.append({
            'data_start': data_start_row,
            'data_end': data_end_row,
            'subtotal_row': subtotal_row,
            'aging_summary_start': aging_summary_start + 1,
            'total_aging_row': total_aging_row,
            'amount_col': amount_col
        })
    
    # Grand totals - only grand total amount
    current_row += 1
    grand_total_row = current_row
    ws.cell(row=current_row, column=1).value = 'GRAND TOTAL'
    ws.cell(row=current_row, column=1).font = Font(bold=True, size=11)
    grand_total_cell = ws.cell(row=current_row, column=2)
    grand_total_cell.value = grand_total_amount
    grand_total_cell.font = Font(bold=True, size=11, underline='single')
    current_row += 2
    
    print(f"\nGrand total: {grand_total_amount}")
    
    # === Add footer ===
    italic_fmt = Font(italic=True)
    bold_fmt = Font(bold=True)
    regular_fmt = Font()
    italic_lines = {
        "For your convenience, payments may be made via the BDO Bills Payment facility:"
    }
    bold_lines = {
        "1. BDO Bills Payment", "a. BDO Mobile Application or BDO Web Page", "b. Over the Counter",
        "NOTE: Please make checks payable to PHILIPPINES FIRST INSURANCE CO., INC",
        "      Payments via LBP and BOC Peso are available only through special arrangement via fund transfer, with an advance copy of the remittance schedule."
    }
    footer_lines = [
        "For your convenience, payments may be made via the BDO Bills Payment facility:",
        "", "1. BDO Bills Payment", "a. BDO Mobile Application or BDO Web Page",
        "   i. Biller: Philippines First Insurance Co., Inc.", "   ii. Reference Number: HO-0001",
        "b. Over the Counter", "   i. Company Name: Philippines First Insurance Co., Inc.",
        "   ii. Subscriber Name: Your Company Name", "   iii. Subscriber Account Number: HO-0001",
        "", "NOTE: Please make checks payable to PHILIPPINES FIRST INSURANCE CO., INC",
        "      Payments via LBP and BOC Peso are available only through special arrangement via fund transfer, with an advance copy of the remittance schedule."
    ]
    for line in footer_lines:
        cell = ws.cell(row=current_row, column=1)
        cell.value = line
        if line in italic_lines:
            cell.font = italic_fmt
        elif line in bold_lines:
            cell.font = bold_fmt
        else:
            cell.font = regular_fmt
        current_row += 1
    
    # === PHASE 2: FORMATTING ONLY ===
    print("\n--- PHASE 2: FORMATTING ---")
    
    # Define the standard accounting number format
    accounting_format = '#,##0.00;[Red](#,##0.00);0.00'
    
    # Format each section
    for idx, section in enumerate(section_positions):
        print(f"\nFormatting section {idx + 1}")
        amount_col = section['amount_col']
        
        if amount_col:
            # Format data rows
            for row in range(section['data_start'], section['data_end'] + 1):
                cell = ws.cell(row=row, column=amount_col)
                if cell.value is not None and isinstance(cell.value, (int, float)):
                    cell.number_format = accounting_format
            
            # Format subtotal
            subtotal_cell = ws.cell(row=section['subtotal_row'], column=amount_col)
            if subtotal_cell.value is not None and isinstance(subtotal_cell.value, (int, float)):
                subtotal_cell.number_format = accounting_format
                subtotal_cell.font = Font(bold=True)
        
        # Format aging summary
        for row in range(section['aging_summary_start'], section['total_aging_row'] + 1):
            value_cell = ws.cell(row=row, column=2)
            if value_cell.value is not None and isinstance(value_cell.value, (int, float)):
                value_cell.number_format = accounting_format
                
                # Apply styles (bold/underline), color is handled by the format string
                label_cell = ws.cell(row=row, column=1)
                if label_cell.value == 'Total':
                    value_cell.font = Font(bold=True, underline='single')
                else:
                    value_cell.font = Font(underline='single')
    
    # Format grand total
    grand_total_cell = ws.cell(row=grand_total_row, column=2)
    if grand_total_cell.value is not None and isinstance(grand_total_cell.value, (int, float)):
        grand_total_cell.number_format = accounting_format
        grand_total_cell.font = Font(bold=True, size=11, underline='single')

    print(f"\n=== FORMATTING COMPLETE ===\n")
    wb.save(file_path)

def extract_soa_reinsurer_cashcall(files):
    """Process SOA for cash call reinsurer"""
    print("=" * 60)
    print("extract_soa_reinsurer_cashcall called")
    print(f"Number of files: {len(files) if files else 0}")
    
    if not files:
        print("ERROR: No files provided")
        return None
    
    if all(f.filename == "" for f in files):
        print("ERROR: All files have empty filenames")
        return None
    
    if len(files) < 2:
        print("ERROR: Less than 2 files provided")
        raise ValueError("Cash Call requires 2 files (bulk and summary)")
    
    temp_dir = tempfile.mkdtemp()
    print(f"Created temp directory: {temp_dir}")
    
    today = datetime.now().strftime("%b %d, %Y")
    print(f"Date string: {today}")
    
    try:
        print("\nCalling process_cashcall...")
        output_dfs = process_cashcall(files)
        print(f"process_cashcall returned {len(output_dfs)} items")
        
        if not output_dfs:
            print("ERROR: No output dataframes created")
            import shutil
            shutil.rmtree(temp_dir, ignore_errors=True)
            return None
        
        excel_files = []
        
        print("\nCreating Excel files...")
        for idx, item in enumerate(output_dfs):
            filename, data, is_merged = item
            
            if is_merged:
                # Merged: data is list of (reinsurer_name, df) tuples
                clean_name = make_filename_safe(filename)
                file_name = f"SOA {clean_name} AS OF {today}.xlsx"
                file_path = os.path.join(temp_dir, file_name)
                
                # Create initial workbook
                initial_df = pd.DataFrame()
                initial_df.to_excel(file_path, index=False, engine='openpyxl')
                
                # Apply merged formatting
                apply_cashcall_formatting_merged(file_path, data)
                
                excel_files.append(file_path)
                print(f"  ✓ Created: {file_name} (Merged - {len(data)} reinsurers)")
            else:
                # Single reinsurer
                clean_name = make_filename_safe(filename)
                file_name = f"SOA {clean_name} AS OF {today}.xlsx"
                file_path = os.path.join(temp_dir, file_name)
                
                data.to_excel(file_path, index=False, engine='openpyxl')
                apply_cashcall_formatting(file_path, filename)
                
                excel_files.append(file_path)
                print(f"  ✓ Created: {file_name}")
        
        zip_filename = f"SOA CASH CALL AS OF {today}.zip"
        zip_path = os.path.join(temp_dir, zip_filename)
        
        print(f"\nCreating ZIP file: {zip_filename}")
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            for file_path in excel_files:
                zipf.write(file_path, os.path.basename(file_path))
                print(f"  Added to ZIP: {os.path.basename(file_path)}")
        
        print(f"\n✓ ZIP file created: {zip_filename}")
        print(f"✓ ZIP path: {zip_path}")
        print(f"✓ Total files in ZIP: {len(excel_files)}")
        
        result = (zip_path, zip_filename, temp_dir)
        print(f"\nReturning: {result}")
        print("=" * 60)
        
        return result
    
    except Exception as e:
        print("=" * 60)
        print(f"ERROR in extract_soa_reinsurer_cashcall: {str(e)}")
        import traceback
        print(traceback.format_exc())
        print("=" * 60)
        
        import shutil
        shutil.rmtree(temp_dir, ignore_errors=True)
        raise e